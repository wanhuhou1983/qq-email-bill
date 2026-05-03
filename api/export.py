"""
Excel导出路由（筛选项与 /search 对齐）
"""
import io
from datetime import date
from decimal import Decimal
from typing import Optional

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from db import get_conn
from api.utils import build_where_clause

router = APIRouter(tags=["export"])


@router.get("/export")
def export_excel(
    bank_code: Optional[str] = Query(None, alias="bank"),
    cardholder: Optional[str] = Query(None),
    min_amount: Optional[float] = Query(None),
    max_amount: Optional[float] = Query(None),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    category: Optional[str] = Query(None),
    bill_cycle: Optional[str] = Query(None),
    trans_type: Optional[str] = Query(None),
    currency: Optional[str] = Query(None),
    card_last4: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None),
    description: Optional[str] = Query(None),
):
    params = {k: v for k, v in locals().items() if k not in ("self",) and v is not None}
    where_sql, values = build_where_clause(params)

    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute(f"""SELECT bank_code, cardholder, card_last4, trans_date, post_date,
                               description, amount, trans_type
                        FROM credit_card_transactions t WHERE 1=1 {where_sql}
                        ORDER BY trans_date DESC, id DESC""", values)
        rows = cur.fetchall()
    finally:
        cur.close(); conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "交易明细"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    for ci, h in enumerate(["银行","持卡人","卡号","交易日期","记账日","交易说明","金额","类型"], 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal="center"); c.border = thin

    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin
            if ci == 7 and isinstance(val, (int, float, Decimal)):
                cell.number_format = "#,##0.00"
                cell.font = Font(color="FF0000" if float(val) > 0 else "00AA00")

    for col in ["A","B","C","D","E","F","G","H"]:
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["F"].width = 40

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    filename = f"credit_card_export_{date.today().isoformat()}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/debit/export")
def debit_export(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    keyword: Optional[str] = Query(None),
):
    conditions = []; values = []
    if start_date:
        conditions.append("AND trans_date >= %s"); values.append(start_date)
    if end_date:
        conditions.append("AND trans_date <= %s"); values.append(end_date)
    if keyword:
        conditions.append("AND (description ILIKE %s OR counterparty_name ILIKE %s)")
        values.append(f"%{keyword}%"); values.append(f"%{keyword}%")

    where = " ".join(conditions)
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute(f"""SELECT trans_date, description, debit, credit, balance,
                               COALESCE(counterparty_name,''), COALESCE(counterparty_bank,''),
                               account_number
                        FROM debit_card_transactions WHERE 1=1 {where}
                        ORDER BY trans_date ASC, id ASC""", values)
        rows = cur.fetchall()
    finally:
        cur.close(); conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "借记卡交易"
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    for ci, h in enumerate(["日期","摘要","支出","收入","余额","交易对手","对方银行","账号"], 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfill
        c.alignment = Alignment(horizontal="center"); c.border = thin

    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin
            if ci in (3,4,5) and isinstance(val, (int, float, Decimal)):
                cell.number_format = "#,##0.00"
                cell.font = Font(color="dc2626" if ci == 3 and val > 0 else "059669" if ci == 4 and val > 0 else "000000")

    for col in ["A","B","C","D","E","F","G","H"]:
        ws.column_dimensions[col].width = 18
    ws.column_dimensions["B"].width = 40

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    filename = f"debit_card_export_{date.today().isoformat()}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})
