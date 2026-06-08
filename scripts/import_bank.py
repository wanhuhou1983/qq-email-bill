# -*- coding: utf-8 -*-
"""
Generic bank Excel → PostgreSQL importer.

Usage: python import_bank.py <bank_code> <xlsx_path>
  bank_code: ABC, CCB, ICBC, etc.

Reads a bank's parsed Excel (from *_parse.py) and inserts into
credit_card_bills + credit_card_transactions.
"""
import os, sys, re
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import psycopg2
import openpyxl
import cardholders

DB_HOST = "localhost"
DB_PORT = 5432
DB_USER = "postgres"
DB_PASSWORD = "InfoHub2026"
DB_NAME = "family_finance"

BANK_NAMES = {
    "ABC": "农业银行", "BOCOM": "交通银行", "CCB": "建设银行",
    "CGB": "广发银行", "CITIC": "中信银行", "CMB": "招商银行",
    "ICBC": "工商银行", "PAB": "平安银行", "CEB": "光大银行",
    "CMBC": "民生银行", "CZB": "浙商银行", "BOC": "中国银行",
    "SPDB": "浦发", "NBC": "宁波银行",
}


def get_conn():
    return psycopg2.connect(
        host=DB_HOST, port=DB_PORT, user=DB_USER,
        password=DB_PASSWORD, database=DB_NAME
    )


def classify_trans_type(desc, direction, card_last4=""):
    d = desc or ""
    if direction == "收入":
        if "还款" in d or "转账还" in d or "银联入账" in d or "卡卡转账" in d:
            return "REPAY"
        if "退款" in d or "退货" in d:
            return "REFUND"
        if "刷卡金" in d:
            return "DEPOSIT"
        return "REPAY"
    else:  # 支出
        if "分期" in d:
            if "手续费" in d or "利息" in d:
                return "INSTALLMENT_INT"
            return "INSTALLMENT_PRIN"
        if "年费" in d or "手续费" in d or "利息" in d or "费用" in d:
            return "FEE"
        return "SPEND"


def parse_date(s):
    """Parse various date formats: YYYY-MM-DD, YYYYMMDD, YYYY/MM/DD"""
    s = str(s).strip().replace('/', '-')
    m = re.match(r'(\d{4})-(\d{2})-(\d{2})', s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.match(r'(\d{4})(\d{2})(\d{2})', s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return None


def import_bank(bank_code, xlsx_path):
    bank_name = BANK_NAMES.get(bank_code, bank_code)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    conn = get_conn()
    conn.autocommit = True
    cur = conn.cursor()
    total_bills = 0
    total_txns = 0
    skipped = 0

    try:
        for sheet_name in wb.sheetnames:
            if sheet_name == "summary" or sheet_name == "汇总":
                continue
            if not sheet_name.startswith("card_"):
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue

            header = [str(c or "").strip() for c in rows[0]]

            # Determine card last4 from sheet name: card_8042 or card_1855
            parts = sheet_name.split("_")
            sheet_card = parts[1] if len(parts) >= 2 else ""

            # Build column map from header
            col_map = {}
            for j, h in enumerate(header):
                if "交易日期" in h:
                    col_map["trans_date"] = j
                elif "记账日期" in h:
                    col_map["post_date"] = j
                elif "卡号" in h and "后" in h or "末" in h or "位" in h:
                    col_map["card_last4"] = j
                elif "方向" in h:
                    col_map["direction"] = j
                elif "交易类型" in h:
                    col_map["txn_type"] = j
                elif "交易描述" in h or "交易摘要" in h:
                    col_map["description"] = j
                elif "金额" in h and "交易" in h:
                    col_map["amount"] = j
                elif "入账金额" in h or "结算金额" in h:
                    col_map["settle_amount"] = j

            if "trans_date" not in col_map or "description" not in col_map:
                continue

            # Collect transactions
            txns = []
            for i in range(1, len(rows)):
                row = rows[i]
                if row is None:
                    continue

                def get(idx):
                    return row[idx] if idx is not None and idx < len(row) else None

                td = parse_date(get(col_map["trans_date"]))
                if td is None:
                    continue

                pd_ = parse_date(get(col_map["post_date"])) if "post_date" in col_map else td
                if pd_ is None:
                    pd_ = td

                # Card: try from column, fallback to sheet name
                card_raw = get(col_map.get("card_last4")) if "card_last4" in col_map else None
                card = str(card_raw).strip() if card_raw else sheet_card
                card = card[-4:] if len(card) >= 4 else card

                # Description
                desc = str(get(col_map["description"]) or "").strip()[:500]

                # Amount - use settle_amount if available, else amount
                amt_col = col_map.get("settle_amount") or col_map.get("amount")
                amt_raw = get(amt_col) if amt_col is not None else None
                if amt_raw is None:
                    continue
                try:
                    amt_str = str(amt_raw).replace(',', '').split('/')[0].strip()
                    amt = float(amt_str)
                except:
                    continue

                # Direction
                direction = str(get(col_map.get("direction")) or "").strip() if "direction" in col_map else ""

                # Check for amount-based direction (negative = income)
                if not direction:
                    direction = "收入" if amt < 0 else "支出"

                # Sign: PG convention - positive=expense, negative=income
                if "收入" in direction:
                    signed_amt = -abs(amt)
                else:
                    signed_amt = abs(amt)

                trans_type = classify_trans_type(desc, direction)

                txns.append({
                    "trans_date": td, "post_date": pd_,
                    "card_last4": card, "description": desc,
                    "amount": signed_amt, "trans_type": trans_type,
                })

            if not txns:
                continue

            # Bill info
            dates = sorted(set(t["trans_date"] for t in txns))
            cycle_start = dates[0]
            cycle_end = dates[-1]
            bill_date = cycle_end
            bill_cycle = cycle_start[:7]
            account_masked = "****" + card

            # Who owns this card?
            bill_cardholder = cardholders.get_cardholder(bank_code, card)

            # Upsert bill
            try:
                cur.execute("""
                    INSERT INTO credit_card_bills
                    (bank_code, bank_name, cardholder, bill_date, bill_cycle,
                     cycle_start, cycle_end, account_masked)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (bank_code, bill_date, account_masked)
                    DO UPDATE SET cycle_start=EXCLUDED.cycle_start,
                                  cycle_end=EXCLUDED.cycle_end,
                                  bill_cycle=EXCLUDED.bill_cycle,
                                  cardholder=EXCLUDED.cardholder,
                                  updated_at=NOW()
                    RETURNING id
                """, (bank_code, bank_name, bill_cardholder, bill_date,
                      bill_cycle, cycle_start, cycle_end, account_masked))
                bill_id = cur.fetchone()[0]
                total_bills += 1
            except Exception as e:
                print(f"  ERROR bill '{sheet_name}': {e}")
                continue

            # Insert transactions
            inserted = 0
            for t in txns:
                try:
                    txn_cardholder = cardholders.get_cardholder(bank_code, t["card_last4"])
                    cur.execute("""
                        INSERT INTO credit_card_transactions
                        (bill_id, bank_code, cardholder, card_last4, account_masked,
                         trans_date, post_date, description, amount, currency,
                         trans_type, source, raw_line_text)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'import_bank', %s)
                        ON CONFLICT (bank_code, trans_date, post_date, card_last4, description, amount)
                        DO NOTHING
                    """, (bill_id, bank_code, txn_cardholder, t["card_last4"],
                          "****" + t["card_last4"] if t["card_last4"] else "",
                          t["trans_date"], t["post_date"], t["description"],
                          t["amount"], "CNY", t["trans_type"],
                          f"{t['trans_date']}|{t['amount']}|{t['description'][:80]}"))
                    if cur.rowcount > 0:
                        inserted += 1
                    else:
                        skipped += 1
                except Exception as e:
                    pass  # skip individual insert errors in autocommit mode

            total_txns += inserted
            print(f"  {sheet_name}: {len(txns)} parsed, {inserted} new, cycle {cycle_start}~{cycle_end}, {bill_cardholder}")

    finally:
        cur.close()
        conn.close()
        wb.close()

    print(f"\n{bank_code} done: {total_bills} bills, {total_txns} inserted, {skipped} dupes skipped")
    return total_bills, total_txns


def main():
    if len(sys.argv) < 3:
        print("Usage: python import_bank.py <bank_code> <xlsx_path>")
        print("  bank_code: ABC, CCB, ICBC, ...")
        sys.exit(1)

    bank_code = sys.argv[1].upper()
    xlsx_path = sys.argv[2]

    if not os.path.exists(xlsx_path):
        print(f"File not found: {xlsx_path}")
        sys.exit(1)

    print(f"Importing {BANK_NAMES.get(bank_code, bank_code)} ({bank_code}) from: {xlsx_path}")
    import_bank(bank_code, xlsx_path)


if __name__ == "__main__":
    main()
