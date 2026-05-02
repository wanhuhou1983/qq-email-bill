"""
光大银行信用卡账单 - 生成Excel
数据来源：用户手动提供的解析后明细（非邮件拉取）
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import date

# ============ 1. 原始数据 ============
data = [
    ['4365', '2026-03-14', '2026-03-14', '生活服务 网银在线 京东商城', 102.80],
    ['0173', '2026-03-20', '2026-03-20', '还款 支付宝（摘要：信用卡还款）', -1.01],
    ['0173', '2026-03-20', '2026-03-20', '还款 支付宝（摘要：信用卡还款）', -2.01],
    ['0173', '2026-03-20', '2026-03-20', '还款 支付宝（摘要：信用卡还款）', -1.01],
    ['0173', '2026-03-20', '2026-03-20', '还款 支付宝（摘要：信用卡还款）', -1.01],
    ['5973', '2026-03-21', '2026-03-21', '还款 支付宝（摘要：信用卡还款）', -1.01],
    ['5973', '2026-03-21', '2026-03-21', '还款 支付宝（摘要：信用卡还款）', -2.01],
    ['5973', '2026-03-21', '2026-03-21', '还款 支付宝（摘要：信用卡还款）', -1.01],
    ['5973', '2026-03-21', '2026-03-21', '还款 支付宝（摘要：信用卡还款）', -1.01],
    ['5973', '2026-03-27', '2026-03-27', '还款 光大信用卡还款 北京', -259.88],
    ['5973', '2026-03-27', '2026-03-27', '还款 吴华辉（付款尾号：7213） 银联入账', -5.00],
    ['5973', '2026-04-03', '2026-04-03', '餐饮美食 支付宝 上海拉扎斯信息科技有限公司', 11.49],
    ['', '2026-04-10', '2026-04-10', '分期 京东分期 第5/12期 本金274.96', 274.96],
]

columns = ['信用卡尾号', '交易日期', '记账日期', '交易说明', '人民币金额']
df = pd.DataFrame(data, columns=columns)
df['交易日期'] = pd.to_datetime(df['交易日期'])
df['记账日期'] = pd.to_datetime(df['记账日期'])
df['人民币金额'] = df['人民币金额'].astype(float)

# ============ 2. 统计 ============
consumption = df[df['人民币金额'] > 0]
repayment = df[df['人民币金额'] < 0]
total_consumption = consumption['人民币金额'].sum()
total_repayment = repayment['人民币金额'].sum()
net_amount = df['人民币金额'].sum()

# 按卡汇总
summary_by_card = df.groupby(df['信用卡尾号'].replace('', '分期/汇总'))['人民币金额'].agg(['sum', 'count']).reset_index()
summary_by_card.columns = ['信用卡尾号', '金额合计', '笔数']

print(f"共 {len(df)} 条记录")
print(f"消费 {len(consumption)} 笔 CNY {total_consumption:.2f}")
print(f"还款/退款 {len(repayment)} 笔 CNY {abs(total_repayment):.2f}")
print(f"净额: CNY {net_amount:.2f}")

# ============ 3. 生成 Excel ============
wb = Workbook()

# --- 样式定义 ---
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='4472C4')
money_font_red = Font(color='FF0000')   # 支出红色
money_font_green = Font(color='008000') # 还款绿色
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ===== Sheet 1: 交易明细 =====
ws1 = wb.active
ws1.title = '交易明细'
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws1.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center' if c_idx <= 4 else 'right')
        if r_idx == 1:
            cell.font = header_font
            cell.fill = header_fill
        elif c_idx == 5 and isinstance(value, (int, float)):
            cell.number_format = '#,##0.00'
            if value > 0:
                cell.font = money_font_red
            elif value < 0:
                cell.font = money_font_green

# 调整列宽
ws1.column_dimensions['A'].width = 12
ws1.column_dimensions['B'].width = 14
ws1.column_dimensions['C'].width = 14
ws1.column_dimensions['D'].width = 55
ws1.column_dimensions['E'].width = 14

# ===== Sheet 2: 账单摘要 =====
ws2 = wb.create_sheet('账单摘要')
summary_data = [
    ['银行名称', '中国光大银行'],
    ['账号', '62265541****5973'],
    ['账单周期', '2026-03-11 ~ 2026-04-10'],
    ['导出时间', date.today().strftime('%Y-%m-%d %H:%M')],
    ['总交易数', len(df)],
    ['消费笔数', len(consumption)],
    ['消费总额', f"¥{total_consumption:.2f}"],
    ['还款/退款笔数', len(repayment)],
    ['还款/退款总额', f"¥{abs(total_repayment):.2f}"],
    ['本期净额', f"¥{net_amount:.2f}"],
    ['', ''],
    ['主要支出项目', ''],
    ['- 京东商城', f"¥{(102.80 + 274.96):.2f}"],
    ['- 饿了么(上海拉扎斯)', '¥11.49'],
]
for r_idx, (label, val) in enumerate(summary_data, 1):
    cell_a = ws2.cell(row=r_idx, column=1, value=label)
    cell_b = ws2.cell(row=r_idx, column=2, value=val)
    cell_a.border = thin_border
    cell_b.border = thin_border
    if label and not label.startswith('-') and label != '':
        cell_a.font = Font(bold=True)
        cell_a.fill = PatternFill('solid', fgColor='D9E2F3')

ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 30

# ===== Sheet 3: 按卡汇总 =====
ws3 = wb.create_sheet('按卡汇总')
headers3 = list(summary_by_card.columns) + ['备注']
for c_idx, h in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=c_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

card_remarks = {'4365': '消费卡', '0173': '小额还款卡', '5973': '主卡', '分期/汇总': '京东分期'}
for r_idx, row in summary_by_card.iterrows():
    for c_idx, col in enumerate(summary_by_card.columns, 1):
        cell = ws3.cell(row=r_idx+2, column=c_idx, value=row[col])
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center' if c_idx == 1 else 'right')
        if c_idx == 2:
            cell.number_format = '#,##0.00'
    remark = card_remarks.get(row['信用卡尾号'], '')
    cell_r = ws3.cell(row=r_idx+2, column=4, value=remark)
    cell_r.border = thin_border

ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 14
ws3.column_dimensions['C'].width = 10
ws3.column_dimensions['D'].width = 16

# ============ 4. 保存 ============
output_path = r'C:\Users\linhu\Desktop\光大银行信用卡消费明细.xlsx'
wb.save(output_path)
print(f"\n已保存: {output_path}")
