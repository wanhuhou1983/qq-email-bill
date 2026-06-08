# -*- coding: utf-8 -*-
"""
Import ICBC parsed Excel data into PostgreSQL credit_card tables.

PG sign convention: amount > 0 = expense, amount < 0 = repayment/refund.
ICBC Excel convention: amount is always positive, direction column has '收入'/'支出'.

Usage: python import_icbc.py [xlsx_path]
"""
import os, sys, re
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import psycopg2
import openpyxl
import cardholders

DB_HOST = "localhost"
DB_PORT = 5432
DB_USER = "postgres"
DB_PASSWORD = "InfoHub2026"
DB_NAME = "family_finance"

BANK_CODE = "ICBC"
BANK_NAME = "工商银行"


def get_conn():
    return psycopg2.connect(
        host=DB_HOST, port=DB_PORT, user=DB_USER,
        password=DB_PASSWORD, database=DB_NAME
    )


def classify_trans_type(desc, direction):
    d = desc or ""
    if direction == "收入":
        if "还款" in d or "转账还" in d or "银联入账" in d:
            return "REPAY"
        if "退款" in d or "退货" in d:
            return "REFUND"
        if "刷卡金" in d:
            return "DEPOSIT"
        return "REPAY"
    else:
        if "分期" in d:
            if "手续费" in d or "利息" in d:
                return "INSTALLMENT_INT"
            return "INSTALLMENT_PRIN"
        if "年费" in d or "手续费" in d or "利息" in d or "费用" in d:
            return "FEE"
        return "SPEND"


def import_icbc(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    conn = get_conn()
    conn.autocommit = True  # Each insert auto-commits
    cur = conn.cursor()
    total_bills = 0
    total_txns = 0
    skipped = 0

    try:
        for sheet_name in wb.sheetnames:
            if sheet_name == "汇总":
                continue
            if not sheet_name.startswith("card_"):
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue

            header = rows[0]
            if header[0] != "交易日期":
                continue

            parts = sheet_name.split("_")
            card_last4 = parts[1] if len(parts) >= 2 else ""

            # Collect transactions
            txns = []
            for i in range(1, len(rows)):
                row = rows[i]
                if row is None or row[0] is None:
                    continue
                td_raw = str(row[0]).strip()
                if not re.match(r'\d{4}-\d{2}-\d{2}', td_raw):
                    continue

                td = td_raw[:10]
                pd_ = str(row[1]).strip()[:10] if row[1] else td
                card = str(row[2] or "").strip()[-4:] or card_last4
                desc = str(row[4] or "").strip()[:500]
                amt_raw = row[5]
                direction = str(row[6] or "").strip()

                try:
                    amt = float(str(amt_raw).replace(',', ''))
                except:
                    continue

                if "收入" in direction:
                    signed_amt = -abs(amt)
                else:
                    signed_amt = abs(amt)

                trans_type = classify_trans_type(desc, direction)
                txns.append({
                    "trans_date": td, "post_date": pd_,
                    "card_last4": card, "description": desc,
                    "amount": signed_amt, "trans_type": trans_type,
                })

            if not txns:
                continue

            dates = sorted(set(t["trans_date"] for t in txns))
            cycle_start = dates[0]
            cycle_end = dates[-1]
            bill_date = cycle_end
            bill_cycle = cycle_start[:7]
            account_masked = "****" + card_last4

            # Upsert bill
            try:
                bill_cardholder = cardholders.get_cardholder(BANK_CODE, card_last4)
                cur.execute("""
                    INSERT INTO credit_card_bills
                    (bank_code, bank_name, cardholder, bill_date, bill_cycle,
                     cycle_start, cycle_end, account_masked)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (bank_code, bill_date, account_masked)
                    DO UPDATE SET cycle_start=EXCLUDED.cycle_start,
                                  cycle_end=EXCLUDED.cycle_end,
                                  bill_cycle=EXCLUDED.bill_cycle,
                                  updated_at=NOW()
                    RETURNING id
                """, (BANK_CODE, BANK_NAME, bill_cardholder, bill_date,
                      bill_cycle, cycle_start, cycle_end, account_masked))
                bill_id = cur.fetchone()[0]
                total_bills += 1
            except Exception as e:
                print(f"  ERROR bill '{sheet_name}': {e}")
                continue

            # Insert transactions
            inserted = 0
            for t in txns:
                try:
                    txn_cardholder = cardholders.get_cardholder(BANK_CODE, t["card_last4"])
                    cur.execute("""
                        INSERT INTO credit_card_transactions
                        (bill_id, bank_code, cardholder, card_last4, account_masked,
                         trans_date, post_date, description, amount, currency,
                         trans_type, source, raw_line_text)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'import_icbc', %s)
                        ON CONFLICT (bank_code, trans_date, post_date, card_last4, description, amount)
                        DO NOTHING
                    """, (bill_id, BANK_CODE, txn_cardholder, t["card_last4"],
                          "****" + t["card_last4"] if t["card_last4"] else "",
                          t["trans_date"], t["post_date"], t["description"],
                          t["amount"], "CNY", t["trans_type"],
                          f"{t['trans_date']}|{t['amount']}|{t['description'][:80]}"))
                    if cur.rowcount > 0:
                        inserted += 1
                    else:
                        skipped += 1
                except Exception as e:
                    print(f"    WARN {t['trans_date']} {t['description'][:30]}: {e}")

            total_txns += inserted
            print(f"  {sheet_name}: {len(txns)} parsed, {inserted} new, cycle {cycle_start}~{cycle_end}")

        print(f"\nDone: {total_bills} bills, {total_txns} inserted, {skipped} duplicates skipped")

    finally:
        cur.close()
        conn.close()
        wb.close()


def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\linhu\Documents\信用卡账单\ICBC_direct_extract.xlsx"
    if not os.path.exists(xlsx):
        print(f"File not found: {xlsx}")
        sys.exit(1)
    print(f"Importing ICBC from: {xlsx}")
    import_icbc(xlsx)


if __name__ == "__main__":
    main()
