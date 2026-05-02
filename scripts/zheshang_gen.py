"""
浙商银行信用卡账单 - 生成Excel
数据来源：用户手动提供的解析后明细（非邮件拉取）
账单周期: 2026-03-07 ~ 2026-04-01 | 卡号末四位: 2171
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import date

# ============ 1. 原始数据 ============
data = [
    ['2026-03-07', '2026-03-07', '支付宝（拼多多平台商户）', 106.62, '2171'],
    ['2026-03-08', '2026-03-08', '京东商城（艺振轩新海南鸡饭连锁）', 0.01, '2171'],
    ['2026-03-09', '2026-03-09', '京东商城（沂蒙山炒鸡深山走地鸡31）', 6.08, '2171'],
    ['2026-03-10', '2026-03-10', '支付宝（张帆）', 5019.00, '2171'],
    ['2026-03-11', '2026-03-11', '财付通（山姆会员商店）', 56.80, '2171'],
    ['2026-03-13', '2026-03-13', '支付宝（小淞园）', 6022.80, '2171'],
    ['2026-03-14', '2026-03-14', '财付通（仁济医院）', 6.00, '2171'],
    ['2026-03-16', '2026-03-16', '翼支付', -9000.00, '2171'],          # 大额还款
    ['2026-03-16', '2026-03-16', '银联信用卡还款（71053055）', -942.23, '2171'],
    ['2026-03-17', '2026-03-17', '京东商城（留夫鸭531）', 3.65, '2171'],
    ['2026-03-17', '2026-03-18', '京东支付（京东快递寄件）', 6.00, '2171'],
    ['2026-03-18', '2026-03-18', '支付宝信用卡还款', -2000.00, '2171'],
    ['2026-03-18', '2026-03-18', '京东支付（臻福楼牛羊肉泡馍东方路）', 6.78, '2171'],
    ['2026-03-18', '2026-03-18', '财付通（博海餐饮浦东图书馆店）', 3.00, '2171'],
    ['2026-03-18', '2026-03-18', '财付通（上海由由物业管理有限公司）', 1.00, '2171'],
    ['2026-03-18', '2026-03-18', '财付通（上海由由物业管理有限公司）', -0.51, '2171'], # 调整
    ['2026-03-19', '2026-03-19', '京东商城（京东商城业务）', 4.78, '2171'],
    ['2026-03-19', '2026-03-19', '京东商城业务（退款调整）', -3.88, '2171'],
    ['2026-03-31', '2026-03-31', '财付通（上海由由物业管理有限公司）', 1.00, '2171'],
    ['2026-03-31', '2026-03-31', '财付通（上海由由物业管理有限公司）', -0.65, '2171'], # 调整
    ['2026-04-01', '2026-04-01', '现金分期12期 利息（第2/12期）', 27.44, '2171'],
    ['2026-04-01', '2026-04-01', '现金分期12期 本金摊还（第2/12期）', 1666.67, '2171'],
]

columns = ['交易日期', '记账日期', '交易摘要', '人民币金额', '卡号末四位']
df = pd.DataFrame(data, columns=columns)
df['交易日期'] = pd.to_datetime(df['交易日期'])
df['记账日期'] = pd.to_datetime(df['记账日期'])
df['人民币金额'] = df['人民币金额'].astype(float)

# ============ 2. 统计 ============
consumption = df[df['人民币金额'] > 0]
repayment = df[df['人民币金额'] < 0]
total_consumption = consumption['人民币金额'].sum()
total_repayment = repayment['人民币金额'].sum()
net_amount = df['人民币金额'].sum()

print(f"共 {len(df)} 条记录")
print(f"消费/支出 {len(consumption)} 笔 CNY {total_consumption:.2f}")
print(f"还款/入账 {len(repayment)} 笔 CNY {abs(total_repayment):.2f}")
print(f"净额: CNY {net_amount:.2f}")

# 按卡汇总 + 分类统计
df['交易类型'] = df['人民币金额'].apply(
    lambda x: '消费' if x > 0 else ('还款' if x <= -100 else '调整/退款'))
summary_by_type = df.groupby('交易类型')['人民币金额'].agg(['sum', 'count']).reset_index()
summary_by_type.columns = ['类型', '金额合计', '笔数']

# ============ 3. 生成 Excel ============
wb = Workbook()

header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='C00000')  # 浙商银行红
money_font_red = Font(color='FF0000')
money_font_green = Font(color='008000')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ===== Sheet 1: 交易明细 =====
ws1 = wb.active
ws1.title = '交易明细'
for r_idx, row in enumerate(dataframe_to_rows(df[['交易日期','记账日期','交易摘要','人民币金额','卡号末四位','交易类型']], index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws1.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center' if c_idx <= 5 else ('right' if c_idx == 4 else 'left'))
        if r_idx == 1:
            cell.font = header_font; cell.fill = header_fill
        elif c_idx == 4 and isinstance(value, (int, float)):
            cell.number_format = '#,##0.00'
            if value > 0: cell.font = money_font_red
            elif value < 0: cell.font = money_font_green

ws1.column_dimensions['A'].width = 14
ws1.column_dimensions['B'].width = 14
ws1.column_dimensions['C'].width = 42
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 12
ws1.column_dimensions['F'].width = 12

# ===== Sheet 2: 账单摘要 =====
ws2 = wb.create_sheet('账单摘要')
info = [
    ['银行名称', '浙商银行'],
    ['卡号末四位', '2171'],
    ['账单周期', '2026-03-07 ~ 2026-04-01'],
    ['账单日', '2026-04-01'],
    ['到期还款日', '2026-04-18'],
    ['导出时间', date.today().strftime('%Y-%m-%d %H:%M')],
    ['总交易数', len(df)],
    ['消费笔数', len(consumption)],
    ['消费总额', f"CNY {total_consumption:,.2f}"],
    ['还款笔数', len(repayment)],
    ['还款总额', f"CNY {abs(total_repayment):,.2f}"],
    ['本期净额', f"CNY {net_amount:,.2f}"],
    ['', ''],
    ['账单应还金额(参考)', 'CNY 12,932.59'],
    ['最低还款额(参考)', 'CNY 2,256.03'],
    ['', ''],
    ['大额支出TOP3', ''],
    ['  支付宝（小淞园）', f'CNY 6,022.80'],
    ['  支付宝（张帆）', f'CNY 5,019.00'],
    ['  现金分期本金(第2期)', f'CNY 1,666.67'],
]
for r_idx, (label, val) in enumerate(info, 1):
    cell_a = ws2.cell(row=r_idx, column=1, value=label)
    cell_b = ws2.cell(row=r_idx, column=2, value=val)
    cell_a.border = thin_border
    cell_b.border = thin_border
    if label and not label.startswith('  ') and label != '':
        cell_a.font = Font(bold=True); cell_a.fill = PatternFill('solid', fgColor='FCE4D6')

ws2.column_dimensions['A'].width = 24
ws2.column_dimensions['B'].width = 26

# ===== Sheet 3: 按类型汇总 =====
ws3 = wb.create_sheet('按类型汇总')
headers3 = list(summary_by_type.columns) + ['备注']
type_remarks = {'消费': '含现金分期本金+利息', '还款': '翼支付/银联/支付宝还款', '调整/退款': '物业费调整等'}
for c_idx, h in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=c_idx, value=h)
    cell.font = header_font; cell.fill = header_fill
    cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
for r_idx, row in summary_by_type.iterrows():
    for c_idx, col in enumerate(summary_by_type.columns, 1):
        cell = ws3.cell(row=r_idx+2, column=c_idx, value=row[col])
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center' if c_idx == 1 else 'right')
        if c_idx == 2: cell.number_format = '#,##0.00'
    remark = type_remarks.get(row['类型'], '')
    ws3.cell(row=r_idx+2, column=4, value=remark).border = thin_border

ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 16
ws3.column_dimensions['C'].width = 10
ws3.column_dimensions['D'].width = 24

# ============ 4. 保存 ============
output_path = r'C:\Users\linhu\Desktop\浙商银行信用卡消费明细.xlsx'
wb.save(output_path)
print(f"\n已保存: {output_path}")
