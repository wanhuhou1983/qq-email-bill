"""
中国银行信用卡账单 → 标准Excel
数据来源：QQ邮箱PDF附件 → MinerU解析
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# ========== 原始数据（从MinerU解析的Markdown提取）==========
# 注意：中行格式特殊——存入/支出是两个独立列（都为正数）
# 统一规则：消费=正(+), 还款/存入=负(-)
raw_data = [
    # [交易日期, 记账日, 卡号末四位, 交易说明, 存入金额, 支出金额]
    ["2026-03-07", "2026-03-08", "0177", "拼多多-费列罗巧克力官方旗舰店CHN", None, 58.20],
    ["2026-03-07", "2026-03-08", "0177", "拼多多-M O V A清洁电器 CHN", None, 526.00],
    ["2026-03-09", "2026-03-10", "0177", "上海蝴枋网络科技有限公司CHN", None, 17.80],
    ["2026-03-14", "2026-03-15", "0177", "网银在线-京东商城CHN", None, 101.30],
    ["2026-03-17", "2026-03-17", "0177", "银联转账（云闪付）", 142.92, None],       # 存入→负数
    ["2026-03-19", "2026-03-20", "0177", "京东平台商户CHN", None, 20.00],
    ["2026-03-26", "2026-03-26", "0177", "支付宝支付科技有限公司", 1000.00, None],     # 存入→负数
    ["2026-03-27", "2026-03-27", "0177", "银联转账（云闪付）", 1000.00, None],         # 存入→负数
    ["2026-03-27", "2026-03-28", "0177", "微信-西北匠牛肉面CHN", None, 10.00],
    ["2026-03-27", "2026-03-27", "0177", "无法足额扣款，请补足账户余额", 103.08, None],   # 存入→负数(系统调整)
    ["2026-04-05", "2026-04-06", "0177", "上海燃气有限公司中国（上海）自由贸易试CHN", None, 234.30],
]

# 转换为标准格式：amount统一符号（消费+, 还款-）
standard_data = []
for row in raw_data:
    trans_date, post_date, card_last4, desc, deposit, expenditure = row
    if deposit is not None:
        amount = -abs(deposit)  # 存入→负数
        trans_type = "DEPOSIT"
    else:
        amount = expenditure   # 支出→正数
        trans_type = "SPEND"
    
    standard_data.append({
        "bank_code": "BOC",
        "cardholder": "吴华辉",
        "trans_date": trans_date,
        "post_date": post_date,
        "card_last4": card_last4,
        "card_type": "数字信用卡白金卡",
        "description": desc,
        "category": "",  # 待分类
        "amount": amount,
        "currency": "CNY",
        "trans_type": trans_type,
        "source": "mineru_pdf",
    })

df = pd.DataFrame(standard_data)
df["trans_date"] = pd.to_datetime(df["trans_date"])
df["post_date"] = pd.to_datetime(df["post_date"])

# ========== 创建Excel（3个Sheet）==========
wb = Workbook()

# --- Sheet 1: 交易明细 ---
ws1 = wb.active
ws1.title = "交易明细"

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="B22222", end_color="B22222", fill_type="solid")  # 中行红
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

headers = ["交易日期", "记账日", "卡号末四位", "卡种", "交易说明", "金额", "币种", "交易类型", "来源"]
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

for r_idx, row in enumerate(df.itertuples(), 2):
    ws1.cell(row=r_idx, column=1, value=row.trans_date.strftime("%Y-%m-%d")).border = thin_border
    ws1.cell(row=r_idx, column=2, value=row.post_date.strftime("%Y-%m-%d")).border = thin_border
    ws1.cell(row=r_idx, column=3, value=row.card_last4).border = thin_border
    ws1.cell(row=r_idx, column=4, value=row.card_type).border = thin_border
    ws1.cell(row=r_idx, column=5, value=row.description).border = thin_border
    amt_cell = ws1.cell(row=r_idx, column=6, value=row.amount)
    amt_cell.border = thin_border
    amt_cell.number_format = '#,##0.00'
    if row.amount < 0:
        amt_cell.font = Font(color="00AA00")  # 绿色=还款
    else:
        amt_cell.font = Font(color="FF0000")  # 红色=支出
    ws1.cell(row=r_idx, column=7, value=row.currency).border = thin_border
    ws1.cell(row=r_idx, column=8, value=row.trans_type).border = thin_border
    ws1.cell(row=r_idx, column=9, value=row.source).border = thin_border

ws1.column_dimensions['A'].width = 12
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 10
ws1.column_dimensions['D'].width = 16
ws1.column_dimensions['E'].width = 50
ws1.column_dimensions['F'].width = 12
ws1.column_dimensions['G'].width = 8
ws1.column_dimensions['H'].width = 12
ws1.column_dimensions['I'].width = 12

# --- Sheet 2: 账单摘要 ---
ws2 = wb.create_sheet("账单摘要")
summary_rows = [
    ["银行", "中国银行"],
    ["银行代码", "BOC"],
    ["持卡人", "吴华辉"],
    ["卡种", "数字信用卡白金卡"],
    ["卡号末四位", "0177"],
    ["账号脱敏", "6259 0755 *** 0177"],
    ["", ""],
    ["账单日", "2026-04-07"],
    ["到期还款日", "2026-04-27"],
    ["账单周期起始", "2026-03-07"],
    ["账单周期结束", "2026-04-07"],
    ["", ""],
    ["上期欠款/存款余额", 2142.92],
    ["本期支出总额(Purchase)", 967.60],
    ["本期存入总额(Payments)", 2246.00],
    ["本期应还(New Balance)", 864.52],
    ["最低还款额(Min Payment)", 86.00],
    ["账单可分期金额", 45935.48],
    ["", ""],
    ["积分余额", "21,580"],
    ["", ""],
    ["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ["数据来源", "QQ邮箱PDF附件 → MinerU解析"],
]
for r_idx, (label, val) in enumerate(summary_rows, 1):
    ws2.cell(row=r_idx, column=1, value=label).font = Font(bold=True) if label else None
    v_cell = ws2.cell(row=r_idx, column=2, value=val)
    if isinstance(val, (int, float)):
        v_cell.number_format = '#,##0.00'
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 40

# --- Sheet 3: 按卡汇总 ---
ws3 = wb.create_sheet("按卡汇总")

spend_total = df[df["amount"] > 0]["amount"].sum()
deposit_total = df[df["amount"] < 0]["amount"].sum()
net_total = spend_total + deposit_total

summary_data = [
    ["卡号末四位", "卡种", "交易笔数", "消费支出(+) ", "存入还款(-)", "净额"],
    ["0177", "数字信用卡白金卡", len(df), spend_total, deposit_total, net_total],
    ["", "", "", "", "", ""],
    ["合计", "", len(df), spend_total, deposit_total, net_total],
]

for r_idx, row_data in enumerate(summary_data, 1):
    for c_idx, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r_idx, column=c_idx, value=val)
        cell.border = thin_border
        if r_idx == 1 or (r_idx == 4 and c_idx == 1):
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.font = header_font
        if isinstance(val, (int, float)) and r_idx > 1:
            cell.number_format = '#,##0.00'

for col in ['A','B','C','D','E','F']:
    ws3.column_dimensions[col].width = 18

output_path = r"C:\Users\linhu\Desktop\中国银行信用卡消费明细.xlsx"
wb.save(output_path)
print(f"OK: {output_path}")
print(f"  {len(df)} records")
print(f"  spend: {spend_total:.2f}")
print(f"  deposit: {deposit_total:.2f}")
print(f"  net: {net_total:.2f}")
