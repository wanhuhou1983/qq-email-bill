"""
民生银行信用卡账单 - 生成Excel
数据来源：用户手动提供的解析后明细（非邮件拉取）
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import date

# ============ 1. 原始数据 ============
data = [
    ['02/04/2026', '02/04/2026', '财付通-京东商城平台商户', 1.00, '2705'],
    ['02/04/2026', '02/04/2026', '财付通-京东商城平台商户', 1.00, '2705'],
    ['02/04/2026', '02/04/2026', '财付通-京东商城平台商户', 1.00, '2705'],
    ['02/04/2026', '02/04/2026', '财付通-京东商城平台商户', 1.00, '7293'],
    ['02/04/2026', '02/04/2026', '财付通-京东商城平台商户', 1.00, '7293'],
    ['02/04/2026', '02/04/2026', '财付通-京东商城平台商户', 1.00, '7293'],
    ['02/04/2026', '02/04/2026', '财付通-京东商城平台商户', 1.00, '7293'],
    ['02/04/2026', '02/04/2026', '财付通-京东商城平台商户', 1.00, '7293'],
    ['02/06/2026', '02/06/2026', '财付通-上海燃气', 2.41, '7293'],
    ['02/06/2026', '02/06/2026', '信用卡消费金入账', -2.40, '7293'],
    ['02/06/2026', '02/06/2026', '支付宝-吴华辉', 500.00, '9927'],
    ['02/06/2026', '02/06/2026', '财付通-上海燃气', 1.00, '7293'],
    ['02/06/2026', '02/06/2026', '信用卡消费金入账', -0.87, '7293'],
    ['02/12/2026', '02/12/2026', '和包支付-中国移动集团', 15.96, '0575'],
    ['02/14/2026', '02/14/2026', '抖音支付-吴华辉', -10.00, '2705'],
    ['02/14/2026', '02/14/2026', '抖音支付-吴华辉', -100.00, '2705'],
    ['02/15/2026', '02/15/2026', '抖音支付-吴华辉', -7.00, '2705'],
    ['02/18/2026', '02/18/2026', '吴华辉/付款尾号:7213/银联入账', -5.00, '2705'],
    ['02/18/2026', '02/18/2026', '美心食品(广州)有限公司', 48.54, '2705'],
    ['02/18/2026', '02/18/2026', '财付通(银联云闪付)', 1.00, '2705'],
    ['02/19/2026', '02/19/2026', '支付宝-中国铁路网络有限公司', 144.00, '2544'],
    ['02/20/2026', '02/20/2026', '财付通(银联云闪付)', 1.00, '2705'],
    ['02/20/2026', '02/20/2026', '财付通-瑞幸咖啡', 1.82, '2544'],
    ['02/20/2026', '02/21/2026', '支付宝', -57.50, '2544'],
    ['02/20/2026', '02/21/2026', '支付宝', -57.50, '2544'],
    ['02/21/2026', '02/21/2026', '财付通-滴滴出行', 24.94, '2544'],
    ['02/22/2026', '02/22/2026', '吴华辉/付款尾号:7213/银联入账', -10.00, '2705'],
    ['02/22/2026', '02/22/2026', '财付通-luckincoffee瑞幸咖啡', 15.80, '2544'],
    ['02/24/2026', '02/24/2026', '自动扣账还款', -1371.73, '7293'],
    ['02/28/2026', '03/01/2026', '财付通-滴滴出行', 17.80, '2544'],
    ['02/28/2026', '03/01/2026', '财付通-滴滴出行', 23.00, '2544'],
    ['03/01/2026', '03/01/2026', '财付通-滴滴出行', 12.70, '2544'],
    ['03/02/2026', '03/02/2026', '支付宝-广东赛壹便利店有限公', 12.24, '2544'],
    ['03/03/2026', '03/03/2026', '京东商城平台商户', 16.43, '2705'],
]

columns = ['交易日期', '记账日期', '交易摘要', '人民币金额', '卡号末四位']
df = pd.DataFrame(data, columns=columns)
df['交易日期'] = pd.to_datetime(df['交易日期'], format='%m/%d/%Y')
df['记账日期'] = pd.to_datetime(df['记账日期'], format='%m/%d/%Y')
df['人民币金额'] = df['人民币金额'].astype(float)
df['交易类型'] = df['人民币金额'].apply(lambda x: '消费' if x > 0 else ('还款/退款' if x < 0 else '其他'))

# ============ 2. 统计 ============
consumption = df[df['人民币金额'] > 0]
repayment = df[df['人民币金额'] < 0]
total_consumption = consumption['人民币金额'].sum()
total_repayment = repayment['人民币金额'].sum()
net_amount = df['人民币金额'].sum()

# 按卡汇总
summary_by_card = df.groupby('卡号末四位').agg(
    消费笔数=('人民币金额', lambda x: (x > 0).sum()),
    消费金额=('人民币金额', lambda x: x[x > 0].sum()),
    还款退款笔数=('人民币金额', lambda x: (x < 0).sum()),
    还款退款金额=('人民币金额', lambda x: abs(x[x < 0].sum()))
).reset_index()
summary_by_card['净额'] = summary_by_card['消费金额'] + summary_by_card['还款退款金额'] * (-1)

print(f"共 {len(df)} 条记录")
print(f"消费 {len(consumption)} 笔 CNY {total_consumption:.2f}")
print(f"还款/退款 {len(repayment)} 笔 CNY {abs(total_repayment):.2f}")
print(f"净额: CNY {net_amount:.2f} (账单应还: 731.64)")

# ============ 3. 生成 Excel ============
wb = Workbook()

header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='1E4E79')  # 民生银行蓝
money_font_red = Font(color='FF0000')
money_font_green = Font(color='008000')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ===== Sheet 1: 交易明细 =====
ws1 = wb.active
ws1.title = '交易明细'
df_display = df[['交易类型', '交易日期', '记账日期', '卡号末四位', '交易摘要', '人民币金额']].copy()
for r_idx, row in enumerate(dataframe_to_rows(df_display, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws1.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center' if c_idx <= 5 else ('right' if c_idx == 6 else 'left'))
        if r_idx == 1:
            cell.font = header_font
            cell.fill = header_fill
        elif c_idx == 6 and isinstance(value, (int, float)):
            cell.number_format = '#,##0.00'
            if value > 0:
                cell.font = money_font_red
            elif value < 0:
                cell.font = money_font_green

ws1.column_dimensions['A'].width = 12
ws1.column_dimensions['B'].width = 13
ws1.column_dimensions['C'].width = 13
ws1.column_dimensions['D'].width = 12
ws1.column_dimensions['E'].width = 32
ws1.column_dimensions['F'].width = 14

# ===== Sheet 2: 账单摘要 =====
ws2 = wb.create_sheet('账单摘要')
summary_data = [
    ['银行名称', '中国民生银行'],
    ['账单周期', '2026年02月04日 ~ 2026年03月03日'],
    ['账单日期', '2026年03月03日'],
    ['导出时间', date.today().strftime('%Y-%m-%d %H:%M')],
    ['总交易数', len(df)],
    ['消费笔数', len(consumption)],
    ['消费总额', f"CNY {total_consumption:.2f}"],
    ['还款/退款笔数', len(repayment)],
    ['还款/退款总额', f"CNY {abs(total_repayment):.2f}"],
    ['本期净额(账单应还)', f"CNY {net_amount:.2f}"],
    ['', ''],
    ['按卡分布（5张卡）', ''],
]
for card in sorted(summary_by_card['卡号末四位']):
    row_data = summary_by_card[summary_by_card['卡号末四位'] == card].iloc[0]
    summary_data.append([f"  卡号 ****{card}", f"消费{int(row_data['消费笔数'])}笔 ¥{row_data['消费金额']:.2f} / 还款{int(row_data['还款退款笔数'])}笔"])

for r_idx, (label, val) in enumerate(summary_data, 1):
    cell_a = ws2.cell(row=r_idx, column=1, value=label)
    cell_b = ws2.cell(row=r_idx, column=2, value=val)
    cell_a.border = thin_border
    cell_b.border = thin_border
    if label and not label.startswith('-') and label != '' and not label.startswith('  '):
        cell_a.font = Font(bold=True)
        cell_a.fill = PatternFill('solid', fgColor='DCE6F1')

ws2.column_dimensions['A'].width = 26
ws2.column_dimensions['B'].width = 40

# ===== Sheet 3: 按卡汇总 =====
ws3 = wb.create_sheet('按卡汇总')
headers3 = list(summary_by_card.columns) + ['备注']
for c_idx, h in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=c_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

card_names = {
    '0575': '移动卡',
    '2544': '主力消费卡(火车/滴滴/瑞幸)',
    '2705': '日常消费卡',
    '7293': '主卡(含自动还款)',
    '9927': '大额转账卡',
}
for r_idx, row in summary_by_card.iterrows():
    for c_idx, col in enumerate(summary_by_card.columns, 1):
        cell = ws3.cell(row=r_idx+2, column=c_idx, value=row[col])
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center' if c_idx == 1 else 'right')
        if c_idx in [2, 3, 4, 5]:
            cell.number_format = '#,##0.00'
    remark = card_names.get(row['卡号末四位'], '')
    cell_r = ws3.cell(row=r_idx+2, column=6, value=remark)
    cell_r.border = thin_border

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws3.column_dimensions[col].width = 16 if col != 'F' else 28

# ============ 4. 保存 ============
output_path = r'C:\Users\linhu\Desktop\民生银行信用卡消费明细.xlsx'
wb.save(output_path)
print(f"\n已保存: {output_path}")
