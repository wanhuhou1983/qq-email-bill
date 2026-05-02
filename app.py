"""
信用卡账单查询系统 - FastAPI 后端
"""
import os
import re
from datetime import date
from decimal import Decimal
from typing import Optional

import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from fastapi import FastAPI, Query, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel

load_dotenv()

app = FastAPI(title="信用卡账单查询", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

import os
_HERE = os.path.dirname(os.path.abspath(__file__))

@app.get("/")
def index():
    return FileResponse(os.path.join(_HERE, "index.html"))

# 数据库连接
# 银行名称映射
BANK_NAMES = {
    "ABC": "农业银行", "BOCOM": "交通银行", "CCB": "建设银行", "CGB": "广发银行",
    "CITIC": "中信银行", "CMB": "招商银行", "ICBC": "工商银行", "PAB": "平安银行",
    "CEB": "光大银行", "CMBC": "民生银行", "CZB": "浙商银行", "BOC": "中国银行",
}

def get_conn():
    return psycopg2.connect(
        host=os.getenv("DB_HOST", "localhost"),
        port=int(os.getenv("DB_PORT", 5432)),
        user=os.getenv("DB_USER", "postgres"),
        password=os.getenv("DB_PASSWORD"),
        database=os.getenv("DB_NAME", "postgres"),
    )


# ============ Pydantic 模型 ============

class TransactionItem(BaseModel):
    id: int
    bank_code: str
    bank_name: str
    cardholder: str
    card_last4: str
    card_type: Optional[str]
    trans_date: date
    post_date: date
    description: str
    category: Optional[str]
    amount: float
    currency: str
    trans_type: str
    is_installment: bool
    source: str

class SearchResult(BaseModel):
    total: int
    sum_spend: float = 0
    sum_repay: float = 0
    transactions: list[TransactionItem]


# ============ 工具函数 ============

def build_whereClause(params: dict) -> tuple[str, list]:
    """拼接 WHERE 子句，返回 (sql_str, values)"""
    conditions = []
    values = []

    if params.get("cardholder"):
        conditions.append("AND cardholder ILIKE %s")
        values.append(f"%{params['cardholder']}%")

    if params.get("min_amount") is not None:
        conditions.append("AND amount >= %s")
        values.append(params["min_amount"])

    if params.get("max_amount") is not None:
        conditions.append("AND amount <= %s")
        values.append(params["max_amount"])

    if params.get("start_date"):
        conditions.append("AND trans_date >= %s")
        values.append(params["start_date"])

    if params.get("end_date"):
        conditions.append("AND trans_date <= %s")
        values.append(params["end_date"])

    if params.get("category"):
        conditions.append("AND category ILIKE %s")
        values.append(f"%{params['category']}%")

    if params.get("bank_code"):
        conditions.append("AND bank_code = %s")
        values.append(params["bank_code"])

    if params.get("card_last4"):
        cards = params["card_last4"].split(",")
        placeholders = ",".join(["%s"] * len(cards))
        conditions.append(f"AND card_last4 IN ({placeholders})")
        values.extend(cards)

    if params.get("trans_type") == "EXPENSE":
        conditions.append("AND amount > 0")
    elif params.get("trans_type") == "INCOME":
        conditions.append("AND amount < 0")
    elif params.get("trans_type"):
        conditions.append("AND trans_type = %s")
        values.append(params["trans_type"])

    if params.get("currency"):
        conditions.append("AND currency = %s")
        values.append(params["currency"])

    if params.get("bill_cycle"):
        conditions.append("AND EXISTS (SELECT 1 FROM credit_card_bills b WHERE b.id = t.bill_id AND b.bill_cycle = %s)")
        values.append(params["bill_cycle"])

    return " ".join(conditions), values


def row_to_dict(row, cols) -> dict:
    result = {}
    for col, val in zip(cols, row):
        if isinstance(val, Decimal):
            result[col] = float(val)
        elif isinstance(val, date):
            result[col] = val.isoformat()
        else:
            result[col] = val
    # 补充 bank_name
    bc = result.get("bank_code")
    if bc:
        result["bank_name"] = BANK_NAMES.get(bc, bc)
    return result


# ============ API 端点 ============

@app.get("/api/health")
def health():
    return {"status": "ok"}


@app.get("/api/search", response_model=SearchResult)
def search(
    cardholder: Optional[str] = Query(None, description="持卡人姓名"),
    min_amount: Optional[float] = Query(None, description="最小金额"),
    max_amount: Optional[float] = Query(None, description="最大金额"),
    start_date: Optional[date] = Query(None, description="开始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    category: Optional[str] = Query(None, description="消费类别关键字"),
    bank_code: Optional[str] = Query(None, description="银行代码"),
    cycle_start: Optional[date] = Query(None, description="交易日期起始"),
    cycle_end: Optional[date] = Query(None, description="交易日期截止"),
    bill_cycle: Optional[str] = Query(None, description="账期(YYYY-MM格式，如2026-04)"),
    trans_type: Optional[str] = Query(None, description="交易类型: EXPENSE/INCOME/SPEND/REPAY..."),
    currency: Optional[str] = Query(None, description="币种: CNY/USD/HKD..."),
    card_last4: Optional[str] = Query(None, description="卡号末4位,多卡逗号分隔"),
    limit: int = Query(100, le=500),
    offset: int = Query(0, ge=0),
):
    """
    条件查询交易记录
    """
    params = {
        "cardholder": cardholder,
        "min_amount": min_amount,
        "max_amount": max_amount,
        "start_date": start_date,
        "end_date": end_date,
        "category": category,
        "bank_code": bank_code,
        "bill_cycle": bill_cycle,
        "trans_type": trans_type,
        "currency": currency,
        "card_last4": card_last4,
    }
    where_sql, values = build_whereClause(params)

    base_cols = [
        "id", "bank_code",
        "cardholder", "card_last4", "card_type",
        "trans_date", "post_date",
        "description", "category",
        "amount", "currency",
        "trans_type", "is_installment", "source",
    ]
    col_str = ", ".join(base_cols)

    conn = get_conn()
    cur = conn.cursor()
    try:
        # 总数
        count_sql = f"SELECT COUNT(*) FROM credit_card_transactions t WHERE 1=1 {where_sql}"
        cur.execute(count_sql, values)
        total = cur.fetchone()[0]

        # 聚合：总支出(amount>0) + 总还款(amount<0)
        sum_sql = f"""
            SELECT COALESCE(SUM(CASE WHEN amount > 0 THEN amount ELSE 0 END), 0),
                   COALESCE(SUM(CASE WHEN amount < 0 THEN ABS(amount) ELSE 0 END), 0)
            FROM credit_card_transactions t WHERE 1=1 {where_sql}
        """
        cur.execute(sum_sql, values)
        sum_spend, sum_repay = cur.fetchone()
        sum_spend = float(sum_spend)
        sum_repay = float(sum_repay)

        # 查询
        sql = f"""
            SELECT {col_str}
            FROM credit_card_transactions t
            WHERE 1=1 {where_sql}
            ORDER BY trans_date DESC, id DESC
            LIMIT %s OFFSET %s
        """
        cur.execute(sql, values + [limit, offset])
        rows = cur.fetchall()

        transactions = [row_to_dict(row, base_cols) for row in rows]

        return SearchResult(total=total, sum_spend=sum_spend, sum_repay=sum_repay, transactions=transactions)
    finally:
        cur.close()
        conn.close()


@app.get("/api/banks")
def get_banks():
    """获取所有银行列表"""
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT DISTINCT bank_code, cardholder, COUNT(*) as cnt
            FROM credit_card_transactions
            GROUP BY bank_code, cardholder
            ORDER BY bank_code
        """)
        rows = cur.fetchall()
        return [{"bank_code": r[0], "cardholder": r[1], "count": r[2]} for r in rows]
    finally:
        cur.close()
        conn.close()


@app.get("/api/categories")
def get_categories():
    """获取所有已存在的类别"""
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT DISTINCT category FROM credit_card_transactions
            WHERE category IS NOT NULL AND category != ''
            ORDER BY category
        """)
        rows = cur.fetchall()
        return [r[0] for r in rows if r[0]]
    finally:
        cur.close()
        conn.close()


@app.get("/api/cardholders")
def get_cardholders():
    """获取所有持卡人"""
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT DISTINCT cardholder FROM credit_card_transactions
            WHERE cardholder IS NOT NULL
            ORDER BY cardholder
        """)
        rows = cur.fetchall()
        return [r[0] for r in rows if r[0]]
    finally:
        cur.close()
        conn.close()


@app.get("/api/cards")
def get_cards(bank_code: Optional[str] = Query(None)):
    """获取指定银行下的卡号末4位列表"""
    conn = get_conn()
    cur = conn.cursor()
    try:
        sql = "SELECT DISTINCT card_last4, COUNT(*) as cnt FROM credit_card_transactions WHERE card_last4 IS NOT NULL AND card_last4 != ''"
        vals = []
        if bank_code:
            sql += " AND bank_code = %s"
            vals.append(bank_code)
        sql += " GROUP BY card_last4 ORDER BY card_last4"
        cur.execute(sql, vals)
        rows = cur.fetchall()
        return [{"card_last4": r[0], "count": r[1]} for r in rows]
    finally:
        cur.close()
        conn.close()


@app.get("/api/currencies")
def get_currencies():
    """获取所有币种"""
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT DISTINCT currency FROM credit_card_transactions
            WHERE currency IS NOT NULL AND currency != ''
            ORDER BY currency
        """)
        rows = cur.fetchall()
        return [r[0] for r in rows if r[0]]
    finally:
        cur.close()
        conn.close()


@app.get("/api/ai-search", response_model=SearchResult)
def ai_search(
    q: str = Query(..., description="自然语言查询"),
    limit: int = Query(100, le=500),
    offset: int = Query(0, ge=0),
):
    """
    AI 模糊查询 - 将自然语言转换为 SQL 并执行
    """
    api_key = os.getenv("DEEPSEEK_API_KEY")
    if not api_key:
        return {"error": "未配置 DEEPSEEK_API_KEY，请检查 .env 文件"}

    import openai

    client = openai.OpenAI(api_key=api_key, base_url="https://api.deepseek.com")

    # 先查表结构，让 AI 知道有哪些字段
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT column_name, data_type
            FROM information_schema.columns
            WHERE table_name = 'credit_card_transactions'
            ORDER BY ordinal_position
        """)
        cols = cur.fetchall()
        schema = "\n".join([f"  {c[0]}: {c[1]}" for c in cols])
    finally:
        cur.close()
        conn.close()

    prompt = f"""你是一个 PostgreSQL 查询生成器。根据用户的自然语言查询，生成 SQL 查询。

表名: credit_card_transactions
列结构:
{schema}

重要规则:
- 金额规则: amount > 0 表示消费/支出, amount < 0 表示还款/存入/退款
- 日期字段: trans_date(交易日), post_date(记账日)
- trans_type 可选值: SPEND, REPAY, REFUND, DEPOSIT, INSTALLMENT_PRIN, INSTALLMENT_INT, FEE, CASH_ADVANCE, ADJUST, OTHER
- 持卡人: cardholder
- 自然语言如"消费"对应 amount > 0，"还款"对应 amount < 0

只输出 SQL，不要其他内容，SQL 必须是有效的 SELECT 语句，LIMIT 最大 500 条。
不要输出任何解释，直接输出 SQL。"""

    response = client.chat.completions.create(
        model="deepseek-chat",
        messages=[
            {"role": "system", "content": prompt},
            {"role": "user", "content": q},
        ],
        temperature=0,
        max_tokens=500,
    )

    sql = response.choices[0].message.content.strip()
    # 移除可能的 markdown 代码块
    sql = re.sub(r"^```sql\s*", "", sql, flags=re.IGNORECASE)
    sql = re.sub(r"^```\s*", "", sql)
    sql = sql.strip().rstrip(";")

    # 简单安全检查 + 建议：生产环境应使用只读数据库用户（GRANT SELECT ONLY）
    # 当前黑名单方式无法防止所有绕过（注释、编码等），仅做为基本防线
    if not sql.upper().startswith("SELECT") or "DROP" in sql.upper() or "DELETE" in sql.upper() or "INSERT" in sql.upper() or "UPDATE" in sql.upper() or "TRUNCATE" in sql.upper():
        return {"error": f"生成的 SQL 包含不允许的语句: {sql[:100]}"}

    conn = get_conn()
    cur = conn.cursor()
    try:
        # 聚合
        agg_sql = f"""
            SELECT COALESCE(SUM(CASE WHEN amount > 0 THEN amount ELSE 0 END), 0),
                   COALESCE(SUM(CASE WHEN amount < 0 THEN ABS(amount) ELSE 0 END), 0)
            FROM ({sql}) AS sub
        """
        cur.execute(agg_sql)
        sum_spend, sum_repay = cur.fetchone()
        sum_spend = float(sum_spend)
        sum_repay = float(sum_repay)

        cur.execute(f"SELECT COUNT(*) FROM ({sql}) AS sub")
        total = cur.fetchone()[0]

        cur.execute(sql + f" LIMIT {limit} OFFSET {offset}")
        rows = cur.fetchall()
        cols = [desc[0] for desc in cur.description]

        transactions = [row_to_dict(row, cols) for row in rows]

        return SearchResult(total=total, sum_spend=sum_spend, sum_repay=sum_repay, transactions=transactions)
    except Exception as e:
        return {"error": f"SQL 执行失败: {str(e)}\n生成的SQL: {sql}"}
    finally:
        cur.close()
        conn.close()


@app.get("/api/export")
def export_excel(
    cardholder: Optional[str] = Query(None),
    min_amount: Optional[float] = Query(None),
    max_amount: Optional[float] = Query(None),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    category: Optional[str] = Query(None),
    bank_code: Optional[str] = Query(None),
    bill_cycle: Optional[str] = Query(None),
    trans_type: Optional[str] = Query(None),
    currency: Optional[str] = Query(None),
    card_last4: Optional[str] = Query(None),
):
    """导出当前筛选结果为 Excel"""
    from fastapi.responses import StreamingResponse
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    params = {k: v for k, v in locals().items() if k != "self" and v is not None}
    # 去掉 category（用户已删除该列，但保留筛选兼容性）
    params.pop("category", None)
    where_sql, values = build_whereClause(params)

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(f"""
            SELECT bank_code, cardholder, card_last4, trans_date, post_date,
                   description, amount, trans_type
            FROM credit_card_transactions t WHERE 1=1 {where_sql}
            ORDER BY trans_date DESC
        """, values)
        rows = cur.fetchall()
        cols = [desc[0] for desc in cur.description]
    finally:
        cur.close()
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "交易明细"

    # 表头
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2563EB")
    thin = Border(*(Side(style="thin"),)*4)
    for i, name in enumerate(["银行", "持卡人", "卡号", "交易日", "记账日", "描述", "金额", "类型"], 1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin
            if c == 7:  # 金额
                cell.number_format = '#,##0.00'
                cell.font = Font(color="FF0000" if val > 0 else "008000")
                cell.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=credit_card_export.xlsx"},
    )


@app.post("/api/import-xls")
async def import_xls(file: UploadFile = File(...)):
    """上传浦发银行XLS账单文件并导入数据库"""
    import tempfile, subprocess, json, shutil

    if not file.filename.endswith(".xls"):
        return {"error": "仅支持 .xls 文件", "inserted": 0}

    # 保存上传文件
    tmp = tempfile.NamedTemporaryFile(suffix=".xls", delete=False)
    try:
        content = await file.read()
        tmp.write(content)
        tmp.close()

        # 直接解析XLS（不用subprocess，避免编码问题）
        import pandas as pd
        try:
            df = pd.read_excel(tmp.name, header=None, engine="xlrd")
        except Exception:
            try:
                df = pd.read_excel(tmp.name, header=None, engine="openpyxl")
            except Exception as e2:
                return {"error": f"解析失败: {e2}", "inserted": 0}

        transactions = []
        for i in range(1, len(df)):
            row = df.iloc[i]
            try:
                td = str(int(row[0])); pd_ = str(int(row[1]))
                desc = str(row[2])[:200]; card = str(int(row[3]))
                amt = float(row[6])
                if len(td) == 8 and len(pd_) == 8:
                    transactions.append({
                        "trans_date": td[:4]+"-"+td[4:6]+"-"+td[6:8],
                        "post_date": pd_[:4]+"-"+pd_[4:6]+"-"+pd_[6:8],
                        "description": desc,
                        "amount": amt,
                        "card_last4": card,
                        "cardholder": "吴华辉",
                    })
            except: pass
        if not transactions:
            return {"error": "未找到交易数据", "inserted": 0}

        # 入库
        conn = get_conn()
        cur = conn.cursor()
        try:
            bill_cycle = transactions[0]["trans_date"][:7]
            bill_date = transactions[-1]["post_date"]

            cur.execute("""
                INSERT INTO credit_card_bills (bank_code,bank_name,cardholder,bill_date,bill_cycle,cycle_start,cycle_end,account_masked)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (bank_code,bill_date,account_masked) DO UPDATE SET updated_at=NOW()
                RETURNING id
            """, ("SPDB","浦发银行","吴华辉",bill_date,bill_cycle,
                  transactions[0]["trans_date"],transactions[-1]["trans_date"],
                  "****"+transactions[0].get("card_last4","")))
            bill_id = cur.fetchone()[0]

            inserted = 0
            for t in transactions:
                try:
                    cur.execute("""
                        INSERT INTO credit_card_transactions (bill_id,bank_code,cardholder,card_last4,account_masked,trans_date,post_date,description,amount,currency,trans_type,source,raw_line_text)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'upload',%s)
                        ON CONFLICT (bank_code,trans_date,post_date,card_last4,description,amount) DO NOTHING
                    """, (bill_id,"SPDB",t["cardholder"],t["card_last4"],"****"+t["card_last4"],
                          t["trans_date"],t["post_date"],t["description"],t["amount"],"CNY",
                          "SPEND" if t["amount"] > 0 else "REPAY",
                          f'{t["trans_date"]}|{t["amount"]}|{t["description"]}'))
                    if cur.rowcount > 0: inserted += 1
                except: pass
            conn.commit()
            return {"inserted": inserted, "total": len(transactions), "filename": file.filename, "bill_cycle": bill_cycle}
        finally:
            cur.close(); conn.close()
    finally:
        os.unlink(tmp.name)


@app.post("/api/refresh-qq")
async def refresh_qq():
    """从QQ邮箱拉取最新账单"""
    import subprocess, json

    banks = [
        "abc","boc","bocom","ccb","ceb","cgb","citic","cmb","cmbc","czb","icbc","pab"
    ]
    py = os.path.join(os.path.dirname(__file__), ".venv", "Scripts", "python.exe")
    loader = os.path.join(os.path.dirname(__file__), "bank-loader", "loader.js")

    total_inserted = 0
    bank_count = 0
    errors = []

    for code in banks:
        r = subprocess.run(["node", loader, code], capture_output=True, text=True, encoding="utf-8", errors="replace")
        output = r.stdout + r.stderr
        # 解析新增数
        for line in output.split("\n"):
            if "新增" in line and "条" in line:
                m = __import__("re").search(r"新增 (\d+) 条", line)
                if m:
                    total_inserted += int(m.group(1))
        if "新增" in output:
            bank_count += 1
        if r.returncode != 0 and not "新增" in output:
            errors.append(code)

    return {
        "inserted": total_inserted,
        "banks": bank_count,
        "errors": errors,
        "message": f"刷新完成，{bank_count}家银行有更新，新增{total_inserted}条" + (f"，失败: {errors}" if errors else "")
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8765)
